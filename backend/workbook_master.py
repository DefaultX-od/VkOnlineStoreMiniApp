import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension

from backend.repositories import *
from backend.models import *

PRODUCT_REPO = ProductRepo()

def check_data_integrity(field_names, obj):
    integrity = True
    for field_name in field_names:
        if obj[field_name] is None:
            integrity = False
            break
    return integrity

def adjust_sheet(ws):
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
    ws.column_dimensions = dim_holder
    return ws

def create_category_groups_template(data : list[CategoryGroup] = None):
    template = openpyxl.Workbook()
    working_sheet = template.active

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Группа Категорий'
        working_sheet['B1'] = 'Скрыто'
        is_hidden.add('B2:B11')
    else:
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Группа Категорий'
        working_sheet['C1'] = 'Скрыто'
        row = 2

        for item in data:
            working_sheet[f'A{row}'] = item.id
            working_sheet[f'B{row}'] = item.name
            working_sheet[f'C{row}'] = 'да' if item.is_hidden == True else 'нет'
            row += 1
            pass
        is_hidden.add(f'C2:C{str(len(data) + 1)}')
        pass
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_product_groups_template(data: list[ProductGroup] = None):
    template = openpyxl.Workbook()
    working_sheet = template.active

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Наименование'
        working_sheet['B1'] = 'Скрыто'
        is_hidden.add('B2:B11')
    else:
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Наименование'
        working_sheet['C1'] = 'Скрыто'
        row = 2

        for group in data:
            working_sheet[f'A{row}'] = group.id
            working_sheet[f'B{row}'] = group.name
            working_sheet[f'C{row}'] = 'да' if group.is_hidden == True else 'нет'
            row += 1
            pass
        is_hidden.add(f'C2:C{row-1}')
        pass
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template


def create_categories_template(data : list[Category] = None):
    template = openpyxl.Workbook()
    working_sheet = template.active
    category_groups = PRODUCT_REPO.get_category_groups()
    escaped_categories = [
        f'"{cg.name}"'
        for cg in category_groups
    ]
    cat_group_validation = DataValidation(
        type='list', formula1=','.join(escaped_categories), showDropDown=False
    )

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Наименование'
        working_sheet['B1'] = 'Группа категорий'
        working_sheet['C1'] = 'ID альбома'
        working_sheet['D1'] = 'Скрыто'
        cat_group_validation.add('B2:B11')
        is_hidden.add('D2:D11')
        pass
    else:
        working_sheet['A1'] = 'ID'
        working_sheet['B1'] = 'Наименование'
        working_sheet['C1'] = 'Группа категорий'
        working_sheet['D1'] = 'ID альбома'
        working_sheet['E1'] = 'Скрыто'

        row = 2
        for category in data:
            working_sheet[f'A{row}'] = category.id
            working_sheet[f'B{row}'] = category.name
            working_sheet[f'C{row}'] = category.group.name
            working_sheet[f'D{row}'] = category.icon
            working_sheet[f'E{row}'] = 'да' if category.is_hidden == True else 'нет'
            row +=1
        cat_group_validation.add(f'C2:C{row-1}')
        is_hidden.add(f'E2:E{row-1}')
    working_sheet.add_data_validation(cat_group_validation)
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_products_template(data : list[Product] = None):
    template = openpyxl.Workbook()
    working_sheet = template.active
    categories_list =[
        f'"{category.name}"'
        for category in PRODUCT_REPO.get_categories()
    ]
    item_groups_list =[
        f'"{group.name}"'
        for group in PRODUCT_REPO.get_product_groups()
    ]

    cat_validation = DataValidation(
        type='list', formula1=','.join(categories_list), showDropDown=False
    )
    item_groups_validation = DataValidation(
        type='list', formula1=','.join(item_groups_list), showDropDown=False
    )

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )
    headers = [category_detail.name for category_detail in PRODUCT_REPO.get_category_details()]
    if data is None:
        working_sheet['A1'] = 'Категория'  # список
        working_sheet['B1'] = 'Группа товаров'  # список
        working_sheet['C1'] = 'Название'
        working_sheet['D1'] = 'Цена'
        working_sheet['E1'] = 'Скидка'
        working_sheet['F1'] = 'Альбом'
        working_sheet['G1'] = 'Скрыто'
        col_letter = ord('H')
        for header in headers:
            working_sheet[f'{chr(col_letter)}1'] = header
            col_letter += 1
        cat_validation.add('A2:A11')
        item_groups_validation.add('B2:B11')
        is_hidden.add('G1:G11')
    else:
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Категория' #список
        working_sheet['C1'] = 'Группа товаров' #список
        working_sheet['D1'] = 'Название'
        working_sheet['E1'] = 'Цена'
        working_sheet['F1'] = 'Скидка'
        working_sheet['G1'] = 'Альбом'
        working_sheet['H1'] = 'Скрыто'
        row = 2

        col_letter = ord('I')
        map_category_details = {}

        for header in headers:
            working_sheet[f'{chr(col_letter)}1'] = header
            map_category_details[header] = chr(col_letter)
            col_letter += 1

        for product in data:

            working_sheet[f'A{row}'] = product.id
            working_sheet[f'B{row}'] = product.category.name
            working_sheet[f'C{row}'] = product.group.name
            working_sheet[f'D{row}'] = product.name
            working_sheet[f'E{row}'] = product.price
            working_sheet[f'F{row}'] = product.discount
            working_sheet[f'G{row}'] = product.album_id
            working_sheet[f'H{row}'] = 'да' if product.is_hidden == True else 'нет'

            details = product.details

            col_letter = ord('I')

            for detail in details:
                working_sheet[f'{map_category_details[detail.category_detail.name]}{row}'] = detail.value
                col_letter += 1

            row += 1
            cat_validation.add(f'B2:B{row - 1}')
            item_groups_validation.add(f'C2:C{row - 1}')
            is_hidden.add(f'H2:H{row-1}')

    working_sheet.add_data_validation(cat_validation)
    working_sheet.add_data_validation(item_groups_validation)
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_drop_points_template(data : list[DropPoint] = None):
    template = openpyxl.Workbook()
    working_sheet = template.active

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Страна'
        working_sheet['B1'] = 'Город'
        working_sheet['C1'] = 'Улица'
        working_sheet['D1'] = 'Дом'
        working_sheet['E1'] = 'Примечание'
        working_sheet['F1'] = 'Скрыто'
        is_hidden.add('F1:F11')
    else:
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Страна'
        working_sheet['C1'] = 'Город'
        working_sheet['D1'] = 'Улица'
        working_sheet['E1'] = 'Дом'
        working_sheet['F1'] = 'Примечание'
        working_sheet['G1'] = 'Скрыто'

        row = 2

        for item in data:
            working_sheet[f'A{row}'] = item.id
            working_sheet[f'B{row}'] = item.country
            working_sheet[f'C{row}'] = item.city
            working_sheet[f'D{row}'] = item.street
            working_sheet[f'E{row}'] = item.building
            working_sheet[f'F{row}'] = item.note
            working_sheet[f'G{row}'] = 'да' if item.is_hidden == True else 'нет'

            row +=1
        is_hidden.add(f'G2:G{row-1}')
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_category_details_template(data: list[CategoryDetail] =None):
    template = openpyxl.Workbook()
    working_sheet = template.active

    is_main_validation = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Наименование'
        working_sheet['B1'] = 'Является главной'
        is_main_validation.add('B2:B11')
        pass
    else:
        row = 2
        working_sheet['A1'] = 'ID'
        working_sheet['B1'] = 'Наименование'
        working_sheet['C1'] = 'Является главной'
        for item in data:
            working_sheet[f'A{row}'] = item.id
            working_sheet[f'B{row}'] = item.name
            working_sheet[f'C{row}'] = 'да' if item.is_main == 1 else 'нет'
            row +=1
            pass
        is_main_validation.add(f'C2:C{row-1}')
    working_sheet.add_data_validation(is_main_validation)
    working_sheet = adjust_sheet(working_sheet)
    return template

    pass

def get_data_from_file(file_name):
    path = f'storage/{file_name}'
    file = openpyxl.load_workbook(path)
    working_sheet = file.active

    data = []
    for row in working_sheet.iter_rows(min_row=1, min_col=1, max_row=working_sheet.max_row, max_col=working_sheet.max_column):
        row_data = [cell.value for cell in row]
        if row_data[0] == None:
            break
        else:
            data.append(row_data)
    return data

def construct_category_groups(data):
    category_groups = []
    category_template = {
        'id': None,
        'name': '',
        'is_hidden': 0
    }

    for i in range(1, len(data)):
        template_copy = category_template.copy()
        template_copy['id'] = data[i][0] if len(data[i]) > 2 else None
        template_copy['name'] = data[i][1] if len(data[i]) > 2 else data[i][0]

        is_hidden = data[i][2] if len(data[i]) > 2 else data[i][1]
        template_copy['is_hidden'] = False if is_hidden == 'нет' else True
        if not check_data_integrity(['name', 'is_hidden'], template_copy):
            return None
        else:
            category_groups.append(template_copy)
    return category_groups

def construct_categories(data):
    category_groups_list = { cg.name: cg.id for cg in PRODUCT_REPO.get_category_groups() }

    category_template = {
        'id' : None,
        'name': '',
        'group': 0,
        'album': '',
        'is_hidden': 0
    }
    categories = []
    for row in data[1:]:
        category_template_copy = category_template.copy()
        category_template_copy['id'] = row[0] if len(row)>4 else None
        category_template_copy['name'] = row[1] if len(row) > 4 else row[0]
        category_template_copy['group'] = category_groups_list[row[2]] if len(row) > 4 else category_groups_list[row[1]]
        category_template_copy['album'] = row[3] if len(row) > 4 else row[2]

        is_hidden = row[4] if len(row) > 4 else row[3]
        category_template_copy['is_hidden'] = 0 if is_hidden == 'нет' else 1
        if not check_data_integrity(['name', 'group', 'is_hidden'], category_template_copy):
            return None
        else:
            categories.append(category_template_copy)

    return categories

def construct_item_groups(data):
    category_template = {
        'id': None,
        'name': '',
        'is_hidden' : 0
    }
    categories = []

    for i in range(1, len(data)):
        template_copy = category_template.copy()
        template_copy['id'] = data[i][0] if len(data[i]) > 2 else None
        template_copy['name'] = data[i][1] if len(data[i]) > 2 else data[i][0]

        is_hidden = data[i][2] if len(data[i]) > 2 else data[i][1]
        template_copy['is_hidden'] = 0 if is_hidden == 'нет' else 1

        if not check_data_integrity(['name', 'is_hidden'], template_copy):
            return None
        else:
            categories.append(template_copy)
    return categories

def construct_items_edit(data):
    item_template = {
        'id': None,
        'category': 0,
        'group': 0,
        'name': '',
        'price': 0,
        'discount': 0,
        'album': '',
        'is_hidden': 0,
        'details': []
    }
    categories_list = { c.name: c.id for c in PRODUCT_REPO.get_categories() }
    category_details_list = { cd.name: cd.id for cd in PRODUCT_REPO.get_category_details() }
    product_groups_list = { pg.name: pg.id for pg in PRODUCT_REPO.get_product_groups() }

    items_main_data = []
    for i in range(1, len(data)):
        item_details = []
        item_template_copy = item_template.copy()
        for j, item_detail in enumerate(item_template_copy):
            if item_detail == 'category':
                item_template_copy[item_detail] = categories_list[data[i][j]]
            elif item_detail == 'group':
                item_template_copy[item_detail] = product_groups_list[data[i][j]]
            elif item_detail == 'is_hidden':
                item_template_copy[item_detail] = 0 if data[i][j] == 'нет' else 1
            elif item_detail == 'details':
                for l in range(j, len(data[i])):
                    if data[i][l] is not None:
                        item_details.append({
                            'id': category_details_list[data[0][l]],
                            'value': data[i][l]
                        })
            else:
                item_template_copy[item_detail]=data[i][j]

        item_template_copy['details'] = item_details
        if not check_data_integrity(['category', 'group', 'name', 'price', 'is_hidden'], item_template_copy):
            return None
        else:
            items_main_data.append(item_template_copy)

    return items_main_data


def construct_items_new(data):
    item_template = {
        'category': 0,
        'group': 0,
        'name': '',
        'price': 0,
        'discount': 0,
        'album': '',
        'is_hidden': 0,
        'details': []
    }
    categories_list = {c.name: c.id for c in PRODUCT_REPO.get_categories()}
    category_details_list = {cd.name: cd.id for cd in PRODUCT_REPO.get_category_details()}
    product_groups_list = {pg.name: pg.id for pg in PRODUCT_REPO.get_product_groups()}

    items_main_data = []
    for i in range(1, len(data)):
        item_details = []
        item_template_copy = item_template.copy()
        for j, item_detail in enumerate(item_template_copy):

            if item_detail == 'category':
                item_template_copy[item_detail] = categories_list[data[i][j]]
            elif item_detail == 'group':
                item_template_copy[item_detail] = product_groups_list[data[i][j]]
            elif item_detail == 'is_hidden':
                item_template_copy[item_detail] = 0 if data[i][j] == 'нет' else 1
            elif item_detail == 'details':
                for l in range(j, len(data[i])):
                    if data[i][l] is not None:
                        item_details.append({
                            'id': category_details_list[data[0][l]],
                            'value': data[i][l]
                        })
            else:
                item_template_copy[item_detail]=data[i][j]

        item_template_copy['details'] = item_details
        if not check_data_integrity(['category', 'group', 'name', 'price', 'is_hidden'], item_template_copy):
            return None
        else:
            items_main_data.append(item_template_copy)

    return items_main_data

def construct_items(action, data):
    if action == 'add':
        return construct_items_new(data)
    else:
        return construct_items_edit(data)

def construct_drop_points(data):
    dp_template = {
        'id' : None,
        'country' : '',
        'city' : '',
        'street': '',
        'building' : '',
        'notes' : '',
        'is_hidden' : 0
    }

    drop_points = []

    for i in range(1, len(data)):
        dp_template_copy = dp_template.copy()
        dp_template_copy['id'] = data[i][0] if len(data[i]) > 6 else None
        dp_template_copy['country'] = data[i][1] if len(data[i]) > 6 else data[i][0]
        dp_template_copy['city'] = data[i][2] if len(data[i]) > 6 else data[i][1]
        dp_template_copy['street'] = data[i][3] if len(data[i]) > 6 else data[i][2]
        dp_template_copy['building'] = data[i][4] if len(data[i]) > 6 else data[i][3]
        dp_template_copy['notes'] = data[i][5] if len(data[i]) > 6 else data[i][4]

        is_hidden = data[i][6] if len(data[i]) > 6 else data[i][5]
        dp_template_copy['is_hidden'] = 0 if is_hidden == 'нет' else 1

        if not check_data_integrity(['country', 'city', 'street', 'building', 'is_hidden'], dp_template_copy):
            return None
        else:
            drop_points.append(dp_template_copy)
    return drop_points

def construct_category_details(data):
    category_detail_template = {
        'id' : None,
        'name' : '',
        'is_main': False
    }
    category_details = []
    for i in range(1, len(data)):
        cd_template_copy = category_detail_template.copy()
        cd_template_copy['id'] = data[i][0] if len(data[i]) > 2 else None
        cd_template_copy['name'] = data[i][1] if len(data[i]) > 2 else data[i][0]
        if len(data[i]) > 2:
            cd_template_copy['is_main'] = True if data[i][2] == 'да' else False
        else:
            cd_template_copy['is_main'] = True if data[i][1] == 'да' else False

        if not check_data_integrity(['name','is_main'], cd_template_copy):
            return None

        else:
            category_details.append(cd_template_copy)
    return category_details

def save_doc(doc, mod):
    doc.save(f'storage/doc_{mod}.xlsx')